import os
import sys
import shutil
import subprocess
from pdf2docx import Converter
import pdfplumber
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from .exceptions import DependencyMissingError, UnsupportedFormatError

from .logger import logger

def check_word_installed():
    if sys.platform == "win32":
        try:
            import winreg
            reg = winreg.ConnectRegistry(None, winreg.HKEY_CLASSES_ROOT)
            key = winreg.OpenKey(reg, "Word.Application")
            winreg.CloseKey(key)
            return True
        except (ImportError, OSError):
            return False
    return False

def check_libreoffice_installed():
    return shutil.which("soffice") is not None

def pdf_to_word(file_path, output_path, progress_cb=None):
    if progress_cb: progress_cb(10, "Initializing conversion...")
    cv = Converter(file_path)
    if progress_cb: progress_cb(50, "Converting to Word (this may take a while)...")
    cv.convert(output_path, start=0, end=None)
    cv.close()
    if progress_cb: progress_cb(100, "Done")
    return output_path

def word_to_pdf(file_path, output_path, progress_cb=None):
    if progress_cb: progress_cb(10, "Starting Word to PDF conversion...")
    
    word_installed = check_word_installed()
    lo_installed = check_libreoffice_installed()
    
    if not word_installed and not lo_installed:
        raise DependencyMissingError("Neither MS Word nor LibreOffice found for conversion.")
        
    if word_installed:
        try:
            from docx2pdf import convert
            if progress_cb: progress_cb(50, "Using MS Word to convert...")
            convert(file_path, output_path)
            if progress_cb: progress_cb(100, "Done")
            return output_path
        except Exception as e:
            if not lo_installed:
                raise e
    
    if lo_installed:
        soffice = shutil.which("soffice")
        output_dir = os.path.dirname(output_path)
        if progress_cb: progress_cb(50, "Using LibreOffice fallback...")
        subprocess.run([soffice, "--headless", "--convert-to", "pdf", file_path, "--outdir", output_dir], check=True)
        
        base = os.path.splitext(os.path.basename(file_path))[0]
        lo_output = os.path.join(output_dir, base + ".pdf")
        if lo_output != output_path and os.path.exists(lo_output):
            if os.path.exists(output_path):
                os.remove(output_path)
            os.rename(lo_output, output_path)
            
    if progress_cb: progress_cb(100, "Done")
    return output_path

def pdf_to_excel(file_path, output_path, progress_cb=None):
    if progress_cb: progress_cb(10, "Extracting tables...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    with pdfplumber.open(file_path) as pdf:
        total = len(pdf.pages)
        sheet_count = 0
        for i, page in enumerate(pdf.pages):
            if progress_cb: progress_cb(10 + int((i/total)*80), f"Processing page {i+1}...")
            tables = page.extract_tables()
            if tables:
                for j, table in enumerate(tables):
                    sheet_count += 1
                    ws = wb.create_sheet(title=f"Page {i+1} Tbl {j+1}")
                    for row_idx, row in enumerate(table):
                        for col_idx, cell in enumerate(row):
                            ws.cell(row=row_idx+1, column=col_idx+1, value=cell)
                            
        if sheet_count == 0:
            ws = wb.create_sheet(title="No Tables Found")
            ws.cell(row=1, column=1, value="No tables were found in the PDF.")
            
    wb.save(output_path)
    if progress_cb: progress_cb(100, "Done")
    return output_path

def excel_to_pdf(file_path, output_path, progress_cb=None):
    if progress_cb: progress_cb(10, "Reading Excel file...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    doc = SimpleDocTemplate(output_path, pagesize=letter)
    elements = []
    
    total = len(wb.sheetnames)
    for i, sheet_name in enumerate(wb.sheetnames):
        if progress_cb: progress_cb(10 + int((i/total)*80), f"Processing sheet {sheet_name}...")
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([str(cell) if cell is not None else "" for cell in row])
        
        if data:
            try:
                t = Table(data)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.grey),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0,0), (-1,0), 12),
                    ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                    ('GRID', (0,0), (-1,-1), 1, colors.black),
                ]))
                elements.append(t)
            except Exception as e:
                logger.warning(f"Failed to process sheet {sheet_name}: {e}")
                
    if not elements:
        raise UnsupportedFormatError("No printable data found in the Excel file.")
        
    if progress_cb: progress_cb(90, "Generating PDF...")
    doc.build(elements)
    if progress_cb: progress_cb(100, "Done")
    return output_path